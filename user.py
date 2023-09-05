# user.py
import os
from openpyxl import load_workbook, Workbook
from math import ceil
from dice import csno
import gspread
from oauth2client.service_account import ServiceAccountCredentials

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4
c_macnt = 5
c_casnk = 6
c_mazk = 7
c_seme = 8
c_uke = 9
c_sudoku = 10
c_sdk_prize = 11
data_range = 'A1:K50'
default_money = 10000
default_lvl = 0
default_cnt = 3
default_csn = 10000
default_maz = 10000
default_mazk = 0
default_cascnt = 0


# if os.path.isfile("userDB.xlsx"):

# else:
#     ws.cell(row=1, column=c_name, value="name")
#     ws.cell(row=1, column=c_id, value="id")
#     ws.cell(row=1, column=c_money, value="money")
#     ws.cell(row=1, column=c_lvl, value="lvl")
#     ws.cell(row=1, column=5, value=csno())
#     ws.cell(row=1, column=6, value=default_csn)
#     ws.cell(row=1, column=7, value=default_maz)
#     wb.save("userDB.xlsx")

def savesuko(_row, table, prize):
    wb, ws = readxls()
    #보드판 문자열 하나로 합치기
    if table != 0:
        for l in range(9):
            table[l] = ','.join(list(map(str,table[l])))
        table = '/'.join(list(map(str,table)))
    else:
        table = str(table)
    ws.cell(_row, c_sudoku, table)
    ws.cell(_row, c_sdk_prize, str(prize))
    wb.save("userDB.xlsx")
    wb.close()
    
def readsuko(_row):
    wb, ws = readxls()
    table = ws.cell(_row, c_sudoku).value
    #문자열 보드판배열로 복구
    if table != '0':
        table = table.split('/')
        for l in range(9):
            table[l] = table[l].split(',')
    prize = ws.cell(_row, c_sdk_prize).value
    return table, prize

def readxls():
    if not os.path.isfile("userDB.xlsx"):
        dataget()
    wb = load_workbook('userDB.xlsx')
    ws = wb.active
    return wb, ws


def cnntgsr():
    scope = [
        'https://spreadsheets.google.com/feeds',
        'https://www.googleapis.com/auth/drive',
    ]
    json_file_name = 'dice-mazin-4b8be2be76c1.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json_file_name, scope)
    gc = gspread.authorize(credentials)
    spreadsheet_url = 'https://docs.google.com/spreadsheets/d/1X-7ek4MVuDNevUjnK_AU0ZMo5FgMfwQSNlw8jOApL9E/edit#gid=0'
    wbg = gc.open_by_url(spreadsheet_url)
    wsg = wbg.worksheet('user')
    return wsg


def dataget():
    wsg = cnntgsr()
    range_list = wsg.range(data_range)
    wb = Workbook()
    ws = wb.active
    for cel in range_list:
        ws.cell(cel.row, cel.col, cel.value)
    wb.save("userDB.xlsx")
    wb.close()


def datasave():
    wsg = cnntgsr()
    cel = []
    cels = []
    wb, ws = readxls()
    cell_range = ws[data_range]
    for idx in cell_range:
        for idy in idx:
            cel.append(idy.value)
        cels.append(cel)
        cel = []
    wsg.update(data_range, cels)
    wb.close()
    # print('save complete')


def checkrow():
    wb, ws = readxls()
    for row in range(3, 50):
        if ws.cell(row, 1).value is None:
            wb.close()
            return row
    wb.close()


def signup(_name, _id):
    wb, ws = readxls()
    _row = checkrow()
    ws.cell(_row, c_name, _name)
    ws.cell(_row, c_id, str(hex(_id)))
    ws.cell(_row, c_money, str(default_money))
    ws.cell(_row, c_lvl, str(default_lvl))
    ws.cell(_row, c_macnt, str(default_cnt))
    ws.cell(_row, c_casnk, str(default_cascnt))
    ws.cell(_row, c_mazk, str(default_mazk))
    ws.cell(_row, c_seme, str(0))
    ws.cell(_row, c_uke, str(0))
    ws.cell(_row, c_sudoku, str(0))
    ws.cell(_row, c_sdk_prize, str(0))
    wb.save("userDB.xlsx")
    wb.close()


def findid(_id):
    wb, ws = readxls()
    for row in range(3, 50):
        if ws.cell(row, c_name).value is not None:
            if ws.cell(row, c_id).value == str(hex(_id)):
                wb.close()
                return row
        else:
            break
    wb.close()
    return None


def edtlvl(_row, _lvl):
    wb, ws = readxls()
    if ws.cell(_row, c_name).value is not None:
        ws.cell(_row, c_lvl, str(_lvl))
        wb.save("userDB.xlsx")
        wb.close()


def edtmny(_row, _money):
    wb, ws = readxls()
    if ws.cell(_row, c_name).value is not None:
        ws.cell(_row, c_money, str(_money))
        wb.save("userDB.xlsx")
        wb.close()


def rdinf(_row):
    wb, ws = readxls()
    if ws.cell(_row, c_name).value is not None:
        money = int(ws.cell(_row, c_money).value)
        level = int(ws.cell(_row, c_lvl).value)
        macnt = int(ws.cell(_row, c_macnt).value)
        casnk = int(ws.cell(_row, c_casnk).value)
        wb.close()
        return money, level, macnt, casnk


def rstdat():
    wb, ws = readxls()
    for row in range(3, 50):
        if ws.cell(row, 1).value is not None:
            ws.cell(row, c_money, str(default_money))
            ws.cell(row, c_lvl, str(default_lvl))
            ws.cell(row, c_macnt, str(default_cnt))
            ws.cell(row, c_casnk, str(default_cascnt))
            ws.cell(row, c_mazk, str(default_mazk))
            ws.cell(row, c_seme, str(0))
            ws.cell(row, c_uke, str(0))
            ws.cell(row, c_sudoku, str(0))
            ws.cell(row, c_sdk_prize, str(0))
        else:
            break
    ws.cell(1, 2, str(csno()))
    ws.cell(1, 4, str(default_csn))
    ws.cell(1, 6, str(default_maz))
    wb.save("userDB.xlsx")
    wb.close()


def calbotlvl():
    wb, ws = readxls()
    _row = checkrow()
    botlvl = 0
    if _row > 3:
        for row in range(3, _row):
            botlvl += int(ws.cell(row, c_lvl).value)
        wb.close()
        return ceil(botlvl/(_row-2))


def csnonum():
    wb, ws = readxls()
    csnon = int(ws.cell(1, 2).value)
    wb.close()
    return csnon


def csnokin():
    wb, ws = readxls()
    csnok = int(ws.cell(1, 4).value)
    wb.close()
    return csnok


def csnokined(mny):
    wb, ws = readxls()
    ws.cell(1, 4, str(mny))
    wb.save("userDB.xlsx")
    wb.close()


def csnorst():
    wb, ws = readxls()
    ws.cell(1, 2, str(csno()))
    # csnokined(10000)
    wb.save("userDB.xlsx")
    wb.close()


def mazinkin():
    wb, ws = readxls()
    mazink = int(ws.cell(1, 6).value)
    wb.close()
    return mazink


def mazinkined(mny):
    wb, ws = readxls()
    ws.cell(1, 6, str(mny))
    wb.save("userDB.xlsx")
    wb.close()


def macnted(row, cnt):
    wb, ws = readxls()
    ws.cell(row, c_macnt, str(cnt))
    wb.save("userDB.xlsx")
    wb.close()


def mazinki(row):
    wb, ws = readxls()
    mazki = int(ws.cell(row, c_mazk).value)
    wb.close()
    return mazki


def mazinkied(_row, kig):
    wb, ws = readxls()
    ws.cell(_row, c_mazk, str(kig))
    wb.save("userDB.xlsx")
    wb.close()


def cascnt(row):
    wb, ws = readxls()
    cscnt = int(ws.cell(row, c_casnk).value)
    ws.cell(row, c_casnk, str(cscnt+1))
    wb.save("userDB.xlsx")
    wb.close()
    return int(cscnt)


def rank():
    wb, ws = readxls()
    userrank = {}
    usernum = checkrow()
    for row in range(3, usernum):
        if ws.cell(row, c_id).value != "0x9e8818f3342007b":
            uslvl = int(ws.cell(row, c_lvl).value)
            usname = ws.cell(row, c_name).value
            usmny = int(ws.cell(row, c_money).value)
            userrank[usname] = uslvl, usmny
    result = sorted(userrank.items(), reverse=True, key=lambda item: item[1])
    wb.close()
    return result


def battlew(row, row2):
    wb, ws = readxls()
    if int(ws.cell(row, c_seme).value) == 0 and int(ws.cell(row2, c_uke).value) == 0:
        ws.cell(row, c_seme, str(row2))
        ws.cell(row2, c_uke, str(row))
        wb.save("userDB.xlsx")
        wb.close()
        return True
    else:
        wb.close()
        return False


def battler(row):
    wb, ws = readxls()
    seme = int(ws.cell(row, c_seme).value)
    uke = int(ws.cell(row, c_uke).value)
    wb.close()
    return seme, uke


def battlee(row, row2):
    wb, ws = readxls()
    ws.cell(row, c_seme, str(0))
    ws.cell(row2, c_uke, str(0))
    wb.save("userDB.xlsx")
    wb.close()


def getname(row):
    wb, ws = readxls()
    usname = ws.cell(row, c_name).value
    wb.close()
    return usname


def ruleIDw(ID):
    wb, ws = readxls()
    ws.cell(1, 9, str(ID))
    wb.save("userDB.xlsx")
    wb.close()

def ruleIDr():
    wb, ws = readxls()
    ruleID = ws.cell(1, 9).value
    wb.close()
    return ruleID